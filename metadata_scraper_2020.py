def printtext():
    global e
    global string
    string = e.get()
    text.insert(INSERT, string)

print("Thank you for using Metadata_Scraper_2020! Please begin by entering the URL you would like to search when prompted."+'\n')

import sys
from googlesearch import search
import os
import time
import requests
import re
import tkinter as tk
from tkinter import ttk, simpledialog, Menu
from tkinter import *
from tkinter import messagebox as mbox
from html.parser import HTMLParser
from bs4 import BeautifulSoup
import ctypes
import getpass
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

#creating Excel workbook
wb = Workbook()

dest_filename = 'URL_list.xlsx'

ws1 = wb.active
ws1.title = "Page Data"

wb.save(filename = 'URL_list.xlsx')

URL_DATA = []

#setting up user input box
root = tk.Tk()
root.withdraw()
root.title('Name')
text = Text(root)
e = Entry(root)
e.pack()
e.focus_set()
b = Button(root,text='okay',command=printtext)
text.pack()
b.pack(side='bottom')

#search prompt (ex. academicaffairs.kennesaw.edu)
URL_INP = simpledialog.askstring(title='URL to Search', prompt='What URL would you like to search?')

#search logic
if URL_INP == None:
    sys.exit("Please come again!")
if URL_INP == '' or URL_INP.find('.') == -1:
    sys.exit("This is not a valid URL. Please restart the program and try something else.")
else:
    print("Searching Google for",URL_INP,"and indexing URLs...")

#appending URLs to system
for url in search(('site:'+URL_INP), tld='com', lang='en'):
    URL_DATA.append(url)

if ('urllib.error.HTTPError: HTTP Error 429: Too Many Requests'):
    time.sleep(2)


print("URLs indexed to be scraped:",URL_DATA,'\n','\n','\n')

time.sleep(2000)

print("Scraping indexed URLs for metadata...",'\n')

#printing data
for url in URL_DATA:
    if url == None:
        continue
    else:
        if url.find('.pdf') == -1:
            response = requests.get(url)
            page = response.content
            soup = BeautifulSoup(page, 'lxml')
            print(url)
            URL_DATA.append(soup.find('div', { 'class' : 'content' }).find('p', recursive=False))
            print("Entry complete",'\n')
        else:
            pass
#saving to Excel
else:
    for col in range(1,2):
        ws1.append(URL_DATA)
    print("LIST COMPLETE")
    root = tk.Tk()
    root.withdraw()
    root.title('Name')
    XLSX_INP = simpledialog.askstring(title="Excel Workbook Name", prompt="What would you like to name the Excel sheet for this data?")
    if XLSX_INP == None:
        print("Error in saving workbook, please restart and retry.")
    else:
        wb.save(filename = XLSX_INP+'.xlsx')
        sys.exit("Thank you for flying Air Sam!")
